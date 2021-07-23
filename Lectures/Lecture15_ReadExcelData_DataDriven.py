import openpyxl
from selenium.webdriver.support.select import Select
from Lecture2_WebDriverManager_CrossBrowser import *

class ExcelUtils:

    def __init__(self):
        self.driver = WebDriverManager.webDriverManager(browser="chrome", headless="false")

    def invokeWebSite(self):
        self.driver.get("https://www.orangehrm.com/orangehrm-30-day-trial/")


    @classmethod
    def getRowCount(cls,file,sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetName)
        return sheet.max_row

    @classmethod
    def getColumnCount(cls,file,sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetName)
        return sheet.max_column

    @classmethod
    def readData(cls,file,sheetName,rowNum,colNum):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetName)
        return sheet.cell(row=rowNum,column=colNum).value

    @classmethod
    def writeData(cls,file,sheetName,rowNum,colNum,data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetName)
        sheet.cell(row=rowNum,column=colNum).value = data
        workbook.save(file)


    def enterDataForm(self,file,sheetname):
        url = self.driver.find_element(By.ID,"Form_submitForm_subdomain")
        firstname = self.driver.find_element(By.ID,"Form_submitForm_FirstName")
        lastName = self.driver.find_element(By.ID,"Form_submitForm_LastName")
        email = self.driver.find_element(By.ID,"Form_submitForm_Email")
        jobTitle = self.driver.find_element(By.ID,"Form_submitForm_JobTitle")

        select = Select(self.driver.find_element(By.ID,"Form_submitForm_Industry"))
        company=select.select_by_visible_text("Computer / Technology - Services / Consultancy")

        select = Select(self.driver.find_element(By.ID,"Form_submitForm_NoOfEmployees"))
        totalEmployee=select.select_by_visible_text("76 - 100")

        companyName = self.driver.find_element(By.ID,"Form_submitForm_CompanyName")
        phoneNumber = self.driver.find_element(By.ID,"Form_submitForm_Contact")

        select = Select(self.driver.find_element(By.ID,"Form_submitForm_Country"))
        country=select.select_by_value("India")

        maxRows = ExcelUtils.getRowCount(file,sheetname)
        maxColumns = ExcelUtils.getColumnCount(file,sheetname)
        print(maxRows,maxColumns)

        for row in range(2,maxRows+1):
            for col in range(1,maxColumns+1):
                value=ExcelUtils.readData(file,sheetname,row,col)
                print(value,end= " ")
            print()


if __name__ == "__main__":
    object = ExcelUtils()
    object.invokeWebSite()
    object.enterDataForm("TestData.xlsx","Registration")
